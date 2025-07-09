# -*- coding=utf-8 -*-
# 导入所有需要的库
import math  # 用于数学计算，如此处的向上取整
import time  # 用于时间相关操作，如程序暂停
import requests  # 用于发送HTTP网络请求
from lxml import etree  # 用于解析HTML/XML文档
from datetime import datetime, timedelta  # 用于处理日期和时间
import random  # 用于生成随机数，如随机延迟和User-Agent
import xlsxwriter  # 用于创建和写入Excel (.xlsx) 文件
import smtplib  # 用于发送电子邮件
from email.mime.text import MIMEText  # 用于创建纯文本或HTML格式的邮件内容
from email.mime.multipart import MIMEMultipart  # 用于创建包含多个部分的邮件（如正文和附件）
import csv  # 用于读写CSV文件，但在此脚本中未被使用
import json  # 用于处理JSON数据，但在此脚本中未被使用
import openpyxl  # 用于读取和写入Excel文件，此处用于加载历史数据
import signal  # 用于处理操作系统信号，如此处的Ctrl+C中断
import sys  # 用于与Python解释器交互，如此处的退出程序

# ------------------------- 配置文件 -------------------------
# 在这里配置邮件发送的相关信息，需要替换成您自己的真实信息
SMTP_SERVER = "smtp.example.com"  # SMTP邮件服务器的地址 (例如: "smtp.qq.com")
SMTP_PORT = 465  # SMTP服务器的SSL端口 (例如: 465)
SENDER_EMAIL = "sender@example.com"  # 发件人的邮箱地址
SENDER_PASSWORD = "your_password"  # 发件人邮箱的授权码或密码 (注意：不是登录密码)
RECEIVER_EMAIL = "receiver@example.com"  # 收件人的邮箱地址

# 全局变量，用于在程序运行期间临时保存已抓取到的所有数据
current_data = []


def signal_handler(signum, frame):
    """
    这是一个信号处理器函数，专门用于响应用户按下的 Ctrl+C (SIGINT信号)。
    当用户中断程序时，它会执行收尾工作，而不是让程序直接崩溃。
    """
    print("\n检测到用户中断，正在保存已抓取的数据...")
    # 检查全局变量 current_data 是否有数据
    if current_data:
        # 定义Excel文件的表头
        head = ['序号', '类型', '名称', '日期', '招标人', '代理机构', '区域', '详情', '项目概况']
        # 生成一个带时间戳的文件名，以 "interrupted_data_" 开头
        output_filename = "interrupted_data_" + datetime.now().strftime("%Y%m%d_%H%M%S")
        # 调用 writer_excel 函数将数据写入Excel文件
        writer_excel(current_data, head, '中标公告', output_filename)
        print(f"已保存 {len(current_data)} 条数据到 {output_filename}.xlsx")
    else:
        # 如果没有抓取到任何数据，就打印提示信息
        print("没有数据需要保存。")
    # 正常退出程序
    sys.exit(0)


# 将我们自定义的 signal_handler 函数注册为 SIGINT 信号的处理器
# 这样当用户按下 Ctrl+C 时，Python就会调用这个函数
signal.signal(signal.SIGINT, signal_handler)


# ------------------------- 数据爬取模块 -------------------------
def get_request_headers(referer=None):
    """
    生成一个模拟浏览器的HTTP请求头。
    这有助于伪装我们的爬虫，使其看起来像一个正常的浏览器访问，从而绕过一些基础的反爬虫机制。
    """
    # 定义一个包含多个常见浏览器User-Agent的列表
    user_agents = [
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/101.0.4951.64 Safari/537.36',
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:100.0) Gecko/20100101 Firefox/100.0',
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/101.0.4951.64 Safari/537.36'
    ]
    # 从列表中随机选择一个User-Agent
    ua = random.choice(user_agents)

    # 构建完整的请求头字典
    headers = {
        "User-Agent": ua,  # 模拟浏览器身份
        "Host": "search.ccgp.gov.cn",  # 目标服务器域名
        "Referer": referer if referer else "http://search.ccgp.gov.cn/",  # 指示请求来源页面
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",  # 接受的内容类型
        "Accept-Encoding": "gzip, deflate",  # 接受的编码格式
        "Accept-Language": "zh-CN,zh;q=0.9",  # 接受的语言
        "Connection": "keep-alive",  # 保持连接
        "Upgrade-Insecure-Requests": "1"
    }
    return headers


def open_url(url, params, refer=None):
    """
    发送HTTP GET请求并获取网页内容。
    这个函数封装了请求头生成、随机延迟和实际的请求发送过程。
    """
    # 获取随机生成的请求头
    headers = get_request_headers(refer)

    # 生成一个2到6秒之间的随机延迟
    delay_seconds = random.randint(2, 6)
    print(f"等待 {delay_seconds} 秒以避免频繁请求...")

    try:
        # 程序暂停指定的秒数，这是友好的爬虫行为，可以减轻目标服务器的压力
        time.sleep(delay_seconds)
    except KeyboardInterrupt:
        # 如果在暂停期间用户按下了Ctrl+C，则捕获异常并重新抛出
        print("用户中断了延迟等待")
        raise

    # 使用requests库发送GET请求，并传递URL、请求头和查询参数
    response = requests.get(url, headers=headers, params=params, allow_redirects=True)
    # 检查HTTP响应状态码，如果不是200（成功），则打印错误信息
    if response.status_code != 200:
        print(f"请求失败: {response.status_code}")
    return response


def crawler_ccgp(sheetdata=[], year='', buyerName=''):
    """
    核心爬虫函数，负责抓取中国政府采购网的招标公告数据。
    """
    global current_data  # 声明我们将要修改全局变量 current_data
    current_data = sheetdata  # 将当前数据列表与全局变量同步

    # 定义目标网站的URL和时间范围
    url = 'http://search.ccgp.gov.cn/bxsearch?'
    curr_date = datetime.now()
    start_date = curr_date - timedelta(days=3)  # 设置抓取时间范围为最近3天
    start_time = start_date.strftime("%Y:%m:%d")
    end_time = curr_date.strftime("%Y:%m:%d")

    # 定义HTTP请求的查询参数，这些参数决定了我们要搜索什么内容
    params = {
        'searchtype': 1,
        'page_index': 1,  # 初始页码为1
        'bidSort': 0,  # 公告类型
        'buyerName': buyerName,  # 采购人名称
        'projectId': '',  # 项目编号
        'pinMu': 0,  # 品目
        'bidType': 0,  # 公告类别
        'dbselect': 'bidx',
        'kw': '公告',  # 搜索关键词
        'start_time': start_time,  # 开始时间
        'end_time': end_time,  # 结束时间
        'timeType': 1,  # 时间类型设置为“近三天”
        'displayZone': '',  # 显示区域
        'zoneId': '45',  # 区域ID，这里硬编码为广西的ID
        'pppStatus': 0,  # PPP项目状态
        'agentName': ''  # 代理机构名称
    }

    try:
        print("开始获取数据...")
        # 发送第一次请求以获取总数据量和第一页内容
        resp = open_url(url, params)
        resp.raise_for_status()  # 如果请求失败（非200状态码），则抛出异常
        html = resp.content.decode('utf-8')  # 将响应内容解码为UTF-8字符串
        tree = etree.HTML(html)  # 使用lxml的etree解析HTML

        try:
            # 使用XPath定位并提取公告总数
            total_text = tree.xpath('/html/body/div[5]/div[1]/div/p[1]/span[2]/text()')
            if not total_text:
                # 如果找不到总数，可能是页面结构变化或没有结果
                print("警告：无法在页面上找到数据总数。可能是没有结果或页面结构已更改。")
                total = 0
            else:
                total = int(total_text[0].strip())
        except (IndexError, ValueError):
            # 如果XPath找不到元素或文本无法转换为整数，则打印错误并返回
            print("未找到数据总数，可能没有匹配的结果。")
            return sheetdata

        print(f"找到 {total} 条数据")

        if total > 0:
            # 计算总页数（每页20条数据）
            pagesize = math.ceil(total / 20)
            print(f"总共 {pagesize} 页数据需要抓取")

            # 循环遍历每一页
            for curr_page in range(1, pagesize + 1):
                print(f"正在抓取第 {curr_page}/{pagesize} 页数据...")

                # 使用XPath定位当前页面的所有公告列表项
                list_items = tree.xpath('/html/body/div[5]/div[2]/div/div/div[1]/ul/li')
                # 遍历每个列表项以提取信息
                for i, li in enumerate(list_items):
                    try:
                        # 提取标题、摘要和信息所在的HTML元素
                        title_element = li.find('a')
                        summary_element = li.find('p')
                        span_element = li.find('span')

                        if title_element is None or summary_element is None or span_element is None:
                            print(f"  - 跳过一条不完整的记录。")
                            continue

                        title = title_element.text.strip()
                        link_href = title_element.get('href')
                        summary = summary_element.text.strip() if summary_element.text else ''

                        # 提取并清理包含日期、采购人等信息的文本
                        info = span_element.xpath('string()').replace(' ', '').replace('\r', '').replace('\n',
                                                                                                         '').replace(
                            '\t', '')

                        # 解析info文本以提取各个字段
                        date_part = info[:10]
                        parts = info[10:].split('|')
                        buyer_part = parts[0].replace('采购人：', '') if len(parts) > 0 else ''
                        agent_part = parts[1].replace('代理机构：', '') if len(parts) > 1 else ''
                        region_part = parts[2] if len(parts) > 2 else ''

                        # 将提取的数据组织成一个列表（一行）
                        row = [
                            len(sheetdata) + 1, '公告', title, date_part,
                            buyer_part, agent_part, region_part, link_href, summary
                        ]
                        # 将该行数据添加到结果列表中
                        sheetdata.append(row)
                        current_data = sheetdata  # 实时更新全局数据
                        print(f"  已获取第 {i + 1} 条数据: {title[:30]}...")

                    except (ValueError, IndexError) as e:
                        # 如果在解析过程中出现错误（如字符串分割失败），则打印错误并跳过此条记录
                        print(f"解析数据时出错，跳过此条记录: {e}")
                        continue
                    except KeyboardInterrupt:
                        # 如果用户中断，则打印提示并重新抛出异常，由外层try-except处理
                        print("用户中断了数据抓取")
                        raise

                # 如果当前不是最后一页，则准备请求下一页
                if curr_page < pagesize:
                    params['page_index'] = curr_page + 1  # 更新页码参数
                    print(f"准备抓取下一页数据...")
                    # 发送请求获取下一页内容
                    resp = open_url(url, params, resp.url)
                    html = resp.content.decode('utf-8')
                    tree = etree.HTML(html)

    except KeyboardInterrupt:
        # 捕获用户中断异常
        print("数据抓取被用户中断")
        raise  # 重新抛出，以便主函数可以捕获
    except Exception as e:
        # 捕获其他所有异常，如网络错误
        print(f"抓取数据时发生错误: {e}")
        return sheetdata

    return sheetdata


# ------------------------- 数据处理模块 -------------------------
def load_existing_data(file_path):
    """
    使用 openpyxl 库从一个已存在的 Excel 文件中加载数据。
    这用于获取历史数据，以便进行去重。
    """
    try:
        # 尝试加载Excel工作簿
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active  # 获取活动工作表
        data = []
        headers = []

        # 读取第一行作为表头
        for cell in ws[1]:
            headers.append(cell.value)

        # 从第二行开始遍历所有行，并将每行数据转换为字典
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):  # 确保不是空行
                data.append(dict(zip(headers, row)))

        return data, headers
    except FileNotFoundError:
        # 如果文件不存在，这是正常情况（首次运行），返回空值
        print(f"历史数据文件 '{file_path}' 未找到，将处理所有抓取到的数据为新数据。")
        return None, None
    except Exception as e:
        # 如果发生其他读取错误，打印错误信息
        print(f"读取文件 '{file_path}' 失败: {e}")
        return None, None


def get_existing_titles(data):
    """
    从加载的历史数据中提取所有公告的标题，并存入一个集合中。
    使用集合（set）可以快速进行成员资格检查（去重）。
    """
    if data is not None:
        # 使用集合推导式高效地提取所有'名称'字段的值
        return {item.get('名称', '') for item in data if item.get('名称')}
    return set()


def filter_duplicates(new_data, existing_titles):
    """
    根据已有的标题集合，过滤掉新抓取数据中的重复项。
    """
    # 列表推导式：只保留那些标题（在row[2]）不在 existing_titles 集合中的数据行
    filtered_data = [row for row in new_data if row[2] not in existing_titles]
    return filtered_data


# ------------------------- 邮件通知模块 -------------------------
def send_email(subject, body):
    """
    发送电子邮件。
    """
    # 创建一个MIMEMultipart对象，这是构建复杂邮件的基础
    msg = MIMEMultipart()
    msg["From"] = SENDER_EMAIL  # 设置发件人
    msg["To"] = RECEIVER_EMAIL  # 设置收件人
    msg["Subject"] = subject  # 设置邮件主题
    # 将HTML格式的正文附加到邮件中
    msg.attach(MIMEText(body, "html"))

    try:
        # 使用SMTP_SSL协议连接到邮件服务器，这提供了加密传输
        with smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT) as server:
            server.login(SENDER_EMAIL, SENDER_PASSWORD)  # 登录邮箱
            server.sendmail(SENDER_EMAIL, RECEIVER_EMAIL, msg.as_string())  # 发送邮件
            print("邮件发送成功!")
    except Exception as e:
        # 捕获所有可能的异常，如认证失败、连接超时等
        print(f"邮件发送失败: {e}")


def generate_email_body(new_data):
    """
    根据新数据生成一个精美的HTML格式的邮件正文。
    """
    if not new_data:
        return None

    # 定义HTML表格的列名
    columns = ['序号', '类型', '名称', '日期', '招标人', '代理机构', '区域', '详情', '项目概况']

    # 手动构建HTML表格字符串
    html_rows = []

    # 创建表头行
    header_row = '<tr>' + ''.join([f'<th>{col}</th>' for col in columns]) + '</tr>'
    html_rows.append(header_row)

    # 为每一行数据创建一个HTML表格行
    for row in new_data:
        # 特别处理详情链接，使其成为可点击的超链接
        row_with_link = list(row)
        row_with_link[7] = f'<a href="{row[7]}" target="_blank">点击查看</a>'
        data_row = '<tr>' + ''.join([f'<td>{cell}</td>' for cell in row_with_link]) + '</tr>'
        html_rows.append(data_row)

    # 将所有行合并成一个完整的HTML表格
    html_table = '<table class="data-table">' + ''.join(html_rows) + '</table>'

    # 定义表格的CSS样式，使其更美观
    style = """
    <style>
        .data-table { width: 100%; border-collapse: collapse; font-family: Arial, sans-serif; font-size: 14px; color: #333; }
        .data-table th, .data-table td { padding: 12px; text-align: left; border-bottom: 1px solid #ddd; }
        .data-table th { background-color: #f8f9fa; font-weight: bold; color: #333; }
        .data-table tr:hover { background-color: #f1f1f1; }
        .data-table a { color: #007bff; text-decoration: none; }
        .data-table a:hover { text-decoration: underline; }
    </style>
    """

    # 构建完整的HTML邮件正文
    body = f"""
    <html>
        <head>{style}</head>
        <body>
            <h3>发现 {len(new_data)} 条新招标公告：</h3>
            {html_table}
            <p>请及时查看邮件内容或访问网站获取详细信息。</p>
        </body>
    </html>
    """
    return body


def writer_excel(data, head=['A1', 'A2', 'A3', 'A4', 'A5', 'A6', 'A7', 'A8'], sheetname='sheet1', filename='DataFile'):
    """
    使用 XlsxWriter 库将数据写入一个新的Excel文件。
    """
    # 创建一个新的Excel工作簿
    workbook = xlsxwriter.Workbook(filename + '.xlsx')
    # 添加一个工作表
    worksheet = workbook.add_worksheet(sheetname)

    row = 0
    col = 0

    # 写入表头
    for cvi, cv in enumerate(head):
        worksheet.write(row, col + cvi, cv)
    row += 1

    # 遍历数据并逐行逐单元格写入
    for rowdata in data:
        for cvindex, cv in enumerate(rowdata):
            worksheet.write(row, col + cvindex, cv)
        row += 1

    # 保存并关闭工作簿
    workbook.close()


# ------------------------- 主程序 -------------------------
def main():
    """
    程序的主入口函数，协调所有模块的执行流程。
    """
    try:
        print("开始执行数据爬取任务...")
        print("提示: 按 Ctrl+C 可以中断程序并保存已抓取的数据")

        # 1. 调用爬虫函数抓取数据
        sheetdata = crawler_ccgp([], str(datetime.now().year), '')

        # 2. 加载历史数据用于去重
        # 注意：这里硬编码了文件名，程序需要一个名为 "existing_data.xlsx" 的文件来读取历史记录
        existing_data, headers = load_existing_data("existing_data.xlsx")
        existing_titles = get_existing_titles(existing_data)

        # 3. 过滤掉重复的数据
        filtered_data = filter_duplicates(sheetdata, existing_titles)

        # 4. 根据是否有新数据，决定后续操作
        head = ['序号', '类型', '名称', '日期', '招标人', '代理机构', '区域', '详情', '项目概况']
        if filtered_data:
            # 如果有新数据
            print(f"发现 {len(filtered_data)} 条新数据，准备发送邮件并保存到Excel文件。")
            # 生成邮件正文
            email_body = generate_email_body(filtered_data)
            # 发送邮件
            send_email("[招标公告更新提醒] 发现新数据", email_body)

            # 生成带时间戳的文件名
            output_filename = "filtered_data_" + datetime.now().strftime("%Y%m%d_%H%M%S")
            # 将新数据写入Excel文件
            writer_excel(filtered_data, head, '中标公告', output_filename)
            print(f"新数据已保存到 {output_filename}.xlsx")
        else:
            # 如果没有新数据
            print("未发现新数据，无需发送邮件。")

        print(f"本次共抓取原始数据条数: {len(sheetdata)}")
        print(f"过滤后新增数据条数: {len(filtered_data)}")
        print("任务完成!")

    except KeyboardInterrupt:
        # 如果在主流程中捕获到用户中断，打印提示
        print("\n程序被用户中断。")
    except Exception as e:
        # 捕获所有其他未预料到的异常
        print(f"程序执行过程中发生未知错误: {e}")


# 当该脚本作为主程序直接运行时，执行main()函数
if __name__ == "__main__":
    main()